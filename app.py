print(">>> starting:", __file__, flush=True)
import sys; print(">>> python:", sys.version, flush=True)

from flask import Flask, render_template, request, redirect, url_for, session, flash, send_from_directory, jsonify
from werkzeug.utils import secure_filename
from openpyxl import Workbook, load_workbook
from datetime import datetime, timezone
from functools import wraps
from collections import deque
import os

# ===== RFID Bridge (Alien TagStream) =====
from rfid_bridge import start_tagstream_in_background, read_tags_since, latest_ts, clear_tags

import gspread
from oauth2client.service_account import ServiceAccountCredentials

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
GOOGLE_CREDENTIALS_JSON = os.path.join(BASE_DIR, "credentials.json")

USER_FILE = os.path.join(BASE_DIR, "users.xlsx")
INSPECTION_FILE = os.path.join(BASE_DIR, "inspection.xlsx")
ACTIVITY_FILE = os.path.join(BASE_DIR, "activity_log.xlsx")

UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

GOOGLE_SHEET_ID = "1NecQqjlRiKmrGBWZUIuJrEwRH4B-ShDJb-PsWni1GE0"

# ---- Flask app ----
app = Flask(__name__)
app.secret_key = 'supersecretkey'
app.config.update(SESSION_COOKIE_HTTPONLY=True, SESSION_COOKIE_SAMESITE="Lax")

# ---- Start RFID TagStream listener ----
# start_tagstream_in_background(host="0.0.0.0", port=4000)

HEADERS_USERS = ["username","password","role"]
HEADERS_INSPECTION = [
    "timestamp","username","tag_id","task",
    "หลอดไฟ","เสาไฟ","สายไฟ","latitude","longitude","maps_link",
    "note","location","photo_url","photo_filename",
    "login_at","logout_at"
]
HEADERS_ACTIVITY = ["timestamp","username","action","task","login_at","logout_at"]

def init_excel_files():
    if not os.path.exists(USER_FILE):
        wb = Workbook(); ws = wb.active
        ws.title = "users"; ws.append(HEADERS_USERS)
        wb.save(USER_FILE)
    if not os.path.exists(INSPECTION_FILE):
        wb = Workbook(); ws = wb.active
        ws.title = "inspection"; ws.append(HEADERS_INSPECTION)
        wb.save(INSPECTION_FILE)
    if not os.path.exists(ACTIVITY_FILE):
        wb = Workbook(); ws = wb.active
        ws.title = "activity"; ws.append(HEADERS_ACTIVITY)
        wb.save(ACTIVITY_FILE)
init_excel_files()
def create_admin():
    from openpyxl import load_workbook

    wb = load_workbook(USER_FILE)
    ws = wb.active

    found = False
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0] == "admin":
            found = True
            break

    if not found:
        ws.append(["admin", "1234", "admin"])
        wb.save(USER_FILE)
        print("✅ Created admin user")

    wb.close()

create_admin()

def _gs_client():
    scope = ["https://spreadsheets.google.com/feeds",
             "https://www.googleapis.com/auth/drive"]
    creds = ServiceAccountCredentials.from_json_keyfile_name(GOOGLE_CREDENTIALS_JSON, scope)
    return gspread.authorize(creds)

def get_or_create_worksheet(sh, title, headers):
    try:
        ws = sh.worksheet(title)
    except gspread.exceptions.WorksheetNotFound:
        ws = sh.add_worksheet(title=title, rows=2000, cols=len(headers))
        if headers:
            ws.append_row(headers)
    return ws

def append_inspection_to_excel(data):
    wb = load_workbook(INSPECTION_FILE); ws = wb.active
    ws.append([
        data["timestamp"], data["username"], data["tag_id"], data["task"],
        data["หลอดไฟ"], data["เสาไฟ"], data["สายไฟ"], data["latitude"], data["longitude"], data["maps_link"],
        data["note"], data["location"], data["photo_url"], data["photo_filename"],
        data["login_at"], data["logout_at"]
    ])
    wb.save(INSPECTION_FILE)

def append_inspection_to_gsheets(data):
    client = _gs_client()
    sh = client.open_by_key(GOOGLE_SHEET_ID)
    ws = get_or_create_worksheet(sh, "inspection", HEADERS_INSPECTION)
    ws.append_row([
        data["timestamp"], data["username"], data["tag_id"], data["task"],
        data["หลอดไฟ"], data["เสาไฟ"], data["สายไฟ"], data["latitude"], data["longitude"], data["maps_link"],
        data["note"], data["location"], data["photo_url"], data["photo_filename"],
        data["login_at"], data["logout_at"]
    ])
    print("✅ Wrote inspection to GSheets:", data["tag_id"])

# ---------- helpers for users ----------
def _read_users():
    try:
        wb = load_workbook(USER_FILE)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            u = "" if row[0] is None else str(row[0])
            p = "" if len(row) < 2 or row[1] is None else str(row[1])
            role = "user"
            if len(row) >= 3 and row[2]:
                role = str(row[2])
            if u != "":
                rows.append((u, p, role))
        wb.close()
        return rows
    except Exception:
        return []

def _user_exists(username: str) -> bool:
    for u, p, role in _read_users():   # ✅ รับ 3 ค่า
        if u == username:
            return True
    return False

def _log_activity(action):
    try:
        wb = load_workbook(ACTIVITY_FILE); ws = wb.active
        ws.append([
            datetime.now().isoformat(timespec="seconds"),
            session.get("user",""),
            action,
            session.get("task","inspection"),
            session.get("login_at",""),
            session.get("logout_at",""),
        ])
        wb.save(ACTIVITY_FILE)
    except Exception as e:
        print("⚠️ activity log failed:", e)

# ---------- simulated tag feed (from Dashboard saves) ----------
SIM_TAG_BUFFER = deque(maxlen=500)

def _now_utc_iso():
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00","Z")

def add_sim_tag(epc: str):
    if not epc:
        return
    SIM_TAG_BUFFER.append({
        "ts": _now_utc_iso(),
        "epc": epc,
        "antenna": "sim",
        "rssi": "-",
        "raw": "manual-save"
    })

def sim_tags_since(since: str | None):
    if not since:
        return list(SIM_TAG_BUFFER)
    return [t for t in SIM_TAG_BUFFER if t["ts"] >= since]

# ---------- global auth guard ----------
PUBLIC_ENDPOINTS = {'login', 'register', 'topic3', 'static'}
@app.before_request
def require_login_globally():
    # อนุญาตหน้า public
    if request.endpoint in PUBLIC_ENDPOINTS:
        return
    # บังคับล็อกอิน
    if not session.get('user'):
        return redirect(url_for('login', next=request.path))

# ---------- RFID API (รวมของจริง + ของจำลอง) ----------
@app.get("/api/tags")
def api_tags():
    since = request.args.get("since")
    # จาก hardware bridge (จริง)
    real_items = read_tags_since(since)
    real_latest = latest_ts()
    # จากการบันทึกใน Dashboard (จำลอง)
    sim_items = sim_tags_since(since)
    sim_latest = sim_items[-1]["ts"] if sim_items else None

    items = []
    if real_items: items.extend(real_items)
    if sim_items: items.extend(sim_items)

    latest = real_latest
    if sim_latest and (latest is None or sim_latest > latest):
        latest = sim_latest
    return jsonify({"items": items, "latest_ts": latest})

@app.post("/api/tags/clear")
def api_tags_clear():
    clear_tags()
    SIM_TAG_BUFFER.clear()
    return jsonify({"ok": True})

# ---------- Static uploads ----------
@app.route('/uploads/<path:filename>')
def uploaded_file(filename):
    return send_from_directory(UPLOAD_DIR, filename)

# ---------- Dashboard ----------
@app.route('/dashboard', methods=['GET','POST'])
def dashboard():
    if 'user' not in session:
        return redirect(url_for('login', next=request.path))

    if request.method == 'POST':
        tag_id = request.form.get('post_id','').strip()

        lamp_val   = (request.form.get('หลอดไฟ') or request.form.get('lamp_head') or "").strip()
        pole_val   = (request.form.get('เสาไฟ')  or request.form.get('pole')      or "").strip()
        wiring_val = (request.form.get('สายไฟ')  or request.form.get('wiring')    or "").strip()

        latitude = request.form.get("latitude","").strip()
        longitude = request.form.get("longitude","").strip()
        note = request.form.get('note','').strip()
        location = request.form.get('location','').strip()

        photo = request.files.get('photo')
        photo_filename = ""
        photo_url = ""
        if photo and photo.filename:
            photo_filename = datetime.now().strftime("%Y%m%d_%H%M%S_") + secure_filename(photo.filename)
            photo.save(os.path.join(UPLOAD_DIR, photo_filename))
            base_url = request.host_url.rstrip('/')
            photo_url = f"{base_url}/uploads/{photo_filename}"

        data = {
            "timestamp": datetime.now().isoformat(timespec="seconds"),
            "username": session['user'],
            "tag_id": tag_id,
            "task": session.get("task","inspection"),
            "หลอดไฟ": lamp_val,
            "เสาไฟ": pole_val,
            "สายไฟ": wiring_val,
            "latitude": latitude,
            "longitude": longitude,
            "maps_link": f"https://www.google.com/maps?q={latitude},{longitude}&z=18" if latitude and longitude else "",
            "note": note,
            "location": location,
            "photo_url": photo_url,
            "photo_filename": photo_filename,
            "login_at": session.get("login_at",""),
            "logout_at": ""
        }
        append_inspection_to_excel(data)
        try:
            if os.path.exists(GOOGLE_CREDENTIALS_JSON):
                append_inspection_to_gsheets(data)
        except Exception as e:
            print("❌ GSheet inspection append failed:", e)

        # push tag ไป Topic 3 แบบจำลอง
        add_sim_tag(tag_id)

        session["last_saved"] = data
        flash("บันทึกงานตรวจเรียบร้อย", "ok")
        return redirect('/dashboard')

    last_saved = session.pop("last_saved", None)
    return render_template('dashboard.html',
                       username=session['user'],
                       task=session.get('task','inspection'),
                       last_saved=last_saved,
                       last_tag_id=(last_saved or {}).get("tag_id",""))


# 🔥 วางตรงนี้เลย
from flask import flash, redirect, url_for

@app.route('/admin')
def admin():
    if session.get("role") != "admin":
        flash("คุณไม่มีสิทธิ์เข้าใช้งาน", "err")
        return redirect(url_for("dashboard"))

    keyword = request.args.get("q", "").lower()

    wb = load_workbook(INSPECTION_FILE)
    ws = wb.active

    data = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):

        if not row:
            continue

        row_list = list(row)

        # 🔍 search
        if keyword:
            text = " ".join([str(x).lower() for x in row_list if x])
            if keyword not in text:
                continue

        # ⭐ บังคับให้มี row_id เสมอ
        row_list.append(i)

        data.append(row_list)

    wb.close()

    return render_template("admin.html", data=data, keyword=keyword)

# ---------- Topic 3 ----------
@app.get("/topic3")
def topic3():
    return render_template("topic3.html")

# ---------- Auth: login/register/logout ----------
@app.route("/login", methods=["GET","POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username","").strip()
        password = request.form.get("password","").strip()

        ok = False
        user_role = "user"

        for u, p, role in _read_users():
            if u == username and p == password:
                ok = True
                user_role = role
                break

        if ok:
            session["user"] = username
            session["role"] = user_role
            session["task"] = "inspection"
            session["login_at"] = datetime.now().isoformat(timespec="seconds")
            session["logout_at"] = ""

            _log_activity("login")
            flash("เข้าสู่ระบบสำเร็จ", "ok")

            # ⭐ admin ไปหน้า admin
            if user_role == "admin":
                return redirect("/dashboard")

            # ⭐ user ไป dashboard
            nxt = request.args.get("next") or url_for("dashboard")
            return redirect(nxt)

        else:
            flash("ชื่อผู้ใช้หรือรหัสผ่านไม่ถูกต้อง", "err")

    return render_template("login.html")

@app.route("/register", methods=["GET","POST"])
def register():
    if request.method == "POST":
        username = request.form.get("username","").strip()
        password = request.form.get("password","").strip()
        confirm  = request.form.get("confirm", None)
        if isinstance(confirm, str):
            confirm = confirm.strip()

        if not username or not password:
            flash("กรุณากรอกชื่อผู้ใช้และรหัสผ่าน", "err")
            return render_template("register.html", username=username)

        if confirm is not None and password != confirm:
            flash("รหัสผ่านไม่ตรงกัน", "err")
            return render_template("register.html", username=username)

        if _user_exists(username):
            flash("มีผู้ใช้นี้อยู่แล้ว", "err")
            return render_template("register.html", username=username)

        wb = load_workbook(USER_FILE); ws = wb.active
        ws.append([username, password, "user"])
        wb.save(USER_FILE)

        flash("สมัครสมาชิกสำเร็จ! เข้าสู่ระบบได้เลย", "ok")
        return redirect(url_for("login"))
    return render_template("register.html")

@app.get("/logout")
def logout():
    session["logout_at"] = datetime.now().isoformat(timespec="seconds")
    _log_activity("logout")
    session.clear()
    flash("ออกจากระบบแล้ว", "ok")
    return redirect(url_for("login"))

# ---------- Root ----------
@app.get("/")
def home():
    return redirect(url_for("dashboard") if session.get('user') else url_for("login"))

@app.route("/admin/delete/<int:row_id>")
def delete_row(row_id):
    if session.get("role") != "admin":
        return "ไม่มีสิทธิ์"

    wb = load_workbook(INSPECTION_FILE)
    ws = wb.active

    ws.delete_rows(row_id)
    wb.save(INSPECTION_FILE)
    wb.close()

    return redirect("/admin")

@app.route("/admin/logs")
def admin_logs():
    if session.get("role") != "admin":
        return "ไม่มีสิทธิ์"

    wb = load_workbook(ACTIVITY_FILE)
    ws = wb.active

    logs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        logs.append(row)

    wb.close()

    return render_template("logs.html", logs=logs)

@app.route("/admin/view/<int:row_id>")
def view_detail(row_id):
    if session.get("role") != "admin":
        return "ไม่มีสิทธิ์"

    wb = load_workbook(INSPECTION_FILE)
    ws = wb.active

    row = list(ws.iter_rows(min_row=row_id, max_row=row_id, values_only=True))[0]

    wb.close()

    return render_template("view.html", row=row)

@app.route("/map")
def map_page():
    wb = load_workbook(INSPECTION_FILE)
    ws = wb.active

    poles = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue

        try:
            poles.append({
                "tag_id": row[2],
                "pole_name": "",  # ยังไม่มีใน Excel เดิม
                "latitude": row[7],
                "longitude": row[8],
                "location": row[10],
            })
        except:
            continue

    wb.close()

    return render_template("map.html", poles=poles)

@app.route("/admin/logs/delete/<int:row_id>")
def delete_log(row_id):
    if session.get("role") != "admin":
        return "ไม่มีสิทธิ์"

    wb = load_workbook(ACTIVITY_FILE)
    ws = wb.active

    ws.delete_rows(row_id)
    wb.save(ACTIVITY_FILE)
    wb.close()

    return redirect("/admin/logs")

@app.post("/api/tags")
def receive_tag():
    data = request.json

    SIM_TAG_BUFFER.append({
        "ts": datetime.now().isoformat(),
        "epc": data.get("epc"),
        "rssi": data.get("rssi"),
        "antenna": "rfid"
    })

    return {"ok": True}

if __name__ == '__main__':
    print("📂 BASE_DIR:", BASE_DIR, flush=True)
    print("🔍 Looking for credentials at:", GOOGLE_CREDENTIALS_JSON, flush=True)
    print("✅ Exists?:", os.path.exists(GOOGLE_CREDENTIALS_JSON), flush=True)
    app.run(host='0.0.0.0', port=5000, debug=True)
